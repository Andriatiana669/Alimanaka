from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import models


from .models import TypeRetardConfig, Retard, Rattrapage
from .serializers import (
    TypeRetardConfigSerializer, RetardSerializer, RetardCreateSerializer,
    RattrapageSerializer, RattrapageCreateSerializer
)


class TypeRetardConfigViewSet(viewsets.ReadOnlyModelViewSet):
    """Configuration des types de retards"""
    queryset = TypeRetardConfig.objects.filter(est_actif=True)
    serializer_class = TypeRetardConfigSerializer
    permission_classes = [permissions.IsAuthenticated]


class RetardViewSet(viewsets.ModelViewSet):
    """Gestion des retards"""
    serializer_class = RetardSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # Admin voit tout
        if user.is_staff or user.is_superuser:
            return Retard.objects.all().select_related(
                'utilisateur', 'type_retard'  # ← Ajouté type_retard
            ).prefetch_related('rattrapages')
        
        # Manager voit les retards de son équipe
        if hasattr(user, 'equipes_gerees') and user.equipes_gerees.exists():
            equipes = set()
            for eq in user.equipes_gerees.all():
                equipes.add(eq.id)
                for sous in eq.sous_equipes.all():
                    equipes.add(sous.id)
                    for sous_sous in sous.sous_equipes.all():
                        equipes.add(sous_sous.id)
            
            for eq in user.equipes_co_gerees.all():
                equipes.add(eq.id)
                for sous in eq.sous_equipes.all():
                    equipes.add(sous.id)
                    for sous_sous in sous.sous_equipes.all():
                        equipes.add(sous_sous.id)
            
            return Retard.objects.filter(
                utilisateur__equipe_id__in=equipes
            ).select_related(
                'utilisateur', 'type_retard'  # ← Ajouté type_retard
            ).prefetch_related('rattrapages')
        
        # Utilisateur normal voit ses retards
        return Retard.objects.filter(utilisateur=user).select_related(
            'utilisateur', 'type_retard'  # ← Ajouté type_retard
        ).prefetch_related('rattrapages')
    
    def get_serializer_class(self):
        if self.action == 'create':
            return RetardCreateSerializer
        return RetardSerializer
    
    def perform_create(self, serializer):
        """Crée un retard pour l'utilisateur spécifié ou l'utilisateur connecté"""
        user = self.request.user
        data = self.request.data
        
        # Si user_id fourni et que l'utilisateur peut gérer d'autres
        target_user_id = data.get('user_id')
        if target_user_id and self._can_manage_user(user, target_user_id):
            from users.models import User
            target_user = User.objects.get(id=target_user_id)
        else:
            target_user = user
        
        serializer.save(utilisateur=target_user)
    
    def _can_manage_user(self, manager, target_user_id):
        """Vérifie si le manager peut gérer l'utilisateur cible"""
        if manager.is_superuser:
            return True
        
        from users.models import User
        try:
            target = User.objects.get(id=target_user_id)
        except User.DoesNotExist:
            return False
        
        # Vérifier si dans les équipes gérées
        equipes_manager = set()
        for eq in manager.equipes_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
        
        for eq in manager.equipes_co_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
        
        return target.equipe_id in equipes_manager
    
    @action(detail=False, methods=['get'])
    def mes_retards(self, request):
        """Liste les retards de l'utilisateur connecté avec filtres"""
        queryset = Retard.objects.filter(utilisateur=request.user).select_related('type_retard')
        
        # Filtre par année
        annee = request.query_params.get('annee')
        if annee:
            queryset = queryset.filter(date__year=annee)
        
        # Filtre par statut
        statut = request.query_params.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def calendrier(self, request):
        annee = int(request.query_params.get('annee', datetime.now().year))
        statut_filter = request.query_params.get('statut')
        
        events = []
        
        # 1. Récupérer les retards (avec select_related pour éviter N+1)
        retard_queryset = self.get_queryset().filter(date__year=annee).select_related('type_retard')
        if statut_filter:
            retard_queryset = retard_queryset.filter(statut=statut_filter)
        
        for retard in retard_queryset:
            events.append(retard.to_calendar_event())
        
        # 2. AJOUTER les jours fériés (du module conges)
        from conges.models import JourFerie
        jours_feries = JourFerie.objects.filter(est_actif=True)
        for jf in jours_feries:
            try:
                date_jf = datetime(annee, jf.mois, jf.jour).date()
                events.append({
                    'id': f"ferie_{jf.id}",
                    'title': f"🎉 {jf.nom}",
                    'start': date_jf.isoformat(),
                    'allDay': True,
                    'color': '#ffeb3b',
                    'type': 'ferie',
                    'isBlocked': True
                })
            except ValueError:
                continue
        
        # 3. AJOUTER les jours exceptionnels
        from conges.models import JourExceptionnel
        jours_exc = JourExceptionnel.objects.filter(annee=annee)
        for je in jours_exc:
            events.append({
                'id': f"exc_{je.id}",
                'title': f"⛔ {je.description or 'Exceptionnel'}",
                'start': je.date.isoformat(),
                'allDay': True,
                'color': '#ff5722' if je.type_jour == 'exceptionnel' else '#4caf50',
                'type': 'exceptionnel',
                'isBlocked': je.type_jour == 'exceptionnel'
            })
        
        # 4. AJOUTER les weekends (en dernier)
        from datetime import timedelta
        current = datetime(annee, 1, 1).date()
        while current.year == annee:
            if current.weekday() >= 5:  # Samedi ou Dimanche
                events.append({
                    'id': f"weekend_{current.isoformat()}",
                    'title': "📅 Weekend",
                    'start': current.isoformat(),
                    'allDay': True,
                    'color': '#e0e0e0',
                    'type': 'weekend',
                    'isBlocked': True
                })
            current += timedelta(days=1)
        
        return Response(events)
    
    @action(detail=True, methods=['post'])
    def ajouter_rattrapage(self, request, pk=None):
        """Ajoute une session de rattrapage pour ce retard"""
        retard = self.get_object()
        
        if retard.statut in ['approuve', 'annule']:
            return Response(
                {'error': 'Ce retard est déjà terminé ou annulé'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = RattrapageCreateSerializer(data=request.data)
        if serializer.is_valid():
            rattrapage = serializer.save(
                retard=retard,
                valide_par=request.user
            )
            
            # Recharger le retard pour avoir les infos mises à jour
            retard.refresh_from_db()
            
            return Response({
                'success': True,
                'message': 'Rattrapage ajouté avec succès',
                'rattrapage': RattrapageSerializer(rattrapage).data,
                'heures_restantes': str(retard.heures_restantes)
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def annuler(self, request, pk=None):
        """Annuler un retard"""
        retard = self.get_object()
        user = request.user
        
        if retard.utilisateur != user and not user.is_staff and not self._can_manage_user(user, retard.utilisateur.id):
            return Response({'error': 'Permission refusée'}, status=403)
        
        if retard.statut == 'annule':
            return Response({'error': 'Déjà annulé'}, status=400)
        
        commentaire = request.data.get('commentaire', '')
        
        retard.statut = 'annule'
        retard.annule_par = user
        retard.date_annulation = timezone.now()
        retard.commentaire_annulation = commentaire
        retard.save()
        
        return Response({
            'success': True,
            'message': 'Retard annulé',
            'annule_par': user.get_display_name(),
            'date_annulation': retard.date_annulation
        })
    
    @action(detail=False, methods=['get'])
    def utilisateurs_gerables(self, request):
        """Liste les utilisateurs pour lesquels le manager peut créer des retards"""
        from users.models import User
        
        user = request.user
        
        if user.is_superuser:
            users = User.objects.filter(is_active=True)
        elif user.is_staff:
            equipes = set()
            for eq in user.equipes_gerees.all():
                equipes.add(eq.id)
                for sous in eq.sous_equipes.all():
                    equipes.add(sous.id)
                    for sous_sous in sous.sous_equipes.all():
                        equipes.add(sous_sous.id)
            
            for eq in user.equipes_co_gerees.all():
                equipes.add(eq.id)
                for sous in eq.sous_equipes.all():
                    equipes.add(sous.id)
                    for sous_sous in sous.sous_equipes.all():
                        equipes.add(sous_sous.id)
            
            users = User.objects.filter(
                Q(equipe_id__in=equipes) | Q(id=user.id),
                is_active=True
            ).distinct()
        else:
            users = User.objects.filter(id=user.id)
        
        data = [{
            'id': u.id,
            'display_name': u.get_display_name(),
            'username': u.username,
            'equipe_nom': u.equipe.nom if u.equipe else '',
        } for u in users]
        
        return Response(data)
    
    
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export Excel de tous les retards (admin uniquement)"""
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({'error': 'Permission refusée'}, status=403)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Retards"
        
        # En-têtes
        headers = ['Utilisateur', 'Date', 'Minutes retard', 'Heures à rattraper', 
                   'Heures restantes', 'Type', 'Statut', 'Justificatif', 'Date création']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        retards = Retard.objects.all().select_related('utilisateur', 'type_retard').order_by('-date')
        
        for retard in retards:
            ws.append([
                retard.utilisateur.get_display_name(),
                retard.date.strftime('%d/%m/%Y'),
                retard.minutes_retard,
                float(retard.heures_a_rattraper),
                float(retard.heures_restantes),
                retard.type_retard.nom if retard.type_retard else '-',
                retard.get_statut_display(),
                retard.motif_retard or '',
                retard.date_creation.strftime('%d/%m/%Y %H:%M')
            ])
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=retards_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def export_mine(self, request):
        """Export Excel des retards de l'utilisateur connecté"""
        user = request.user
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mes Retards"
        
        # En-têtes avec infos utilisateur
        ws.append([f"Export des retards - {user.get_display_name()}"])
        ws.append([f"Heures totales à rattraper: {user.retards.filter(statut__in=['en_attente', 'en_cours']).aggregate(total=models.Sum('heures_restantes'))['total'] or 0}h"])
        ws.append([])
        
        # En-têtes du tableau
        headers = ['Date', 'Minutes retard', 'Heures à rattraper', 'Heures restantes', 
                   'Type', 'Statut', 'Justificatif']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Données
        retards = Retard.objects.filter(utilisateur=user).select_related('type_retard').order_by('-date')
        
        for retard in retards:
            ws.append([
                retard.date.strftime('%d/%m/%Y'),
                retard.minutes_retard,
                float(retard.heures_a_rattraper),
                float(retard.heures_restantes),
                retard.type_retard.nom if retard.type_retard else '-',
                retard.get_statut_display(),
                retard.motif_retard or ''
            ])
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=mes_retards_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response


class RattrapageViewSet(viewsets.ReadOnlyModelViewSet):
    """Consultation des rattrapages"""
    serializer_class = RattrapageSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        if user.is_staff or user.is_superuser:
            return Rattrapage.objects.all().select_related('retard__type_retard', 'valide_par')
        
        # Ne voir que ses propres rattrapages
        return Rattrapage.objects.filter(
            retard__utilisateur=user
        ).select_related('retard__type_retard', 'valide_par')