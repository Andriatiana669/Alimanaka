from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from datetime import datetime
from decimal import Decimal  # ← AJOUTÉ

from .models import TypeCongeConfig, JourFerie, JourExceptionnel, CongeAnnuel, Conge
from .serializers import (
    TypeCongeConfigSerializer, JourFerieSerializer,
    JourExceptionnelSerializer, CongeAnnuelSerializer,
    CongeSerializer, CongeCreateSerializer
)


class TypeCongeConfigViewSet(viewsets.ReadOnlyModelViewSet):
    """Configuration des types de congés (lecture seule pour les utilisateurs)"""
    queryset = TypeCongeConfig.objects.all()
    serializer_class = TypeCongeConfigSerializer
    permission_classes = [permissions.IsAuthenticated]


class JourFerieViewSet(viewsets.ReadOnlyModelViewSet):
    """Jours fériés"""
    queryset = JourFerie.objects.filter(est_actif=True)
    serializer_class = JourFerieSerializer
    permission_classes = [permissions.IsAuthenticated]


class JourExceptionnelViewSet(viewsets.ReadOnlyModelViewSet):
    """Jours exceptionnels et non-exceptionnels"""
    serializer_class = JourExceptionnelSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        annee = self.request.query_params.get('annee', datetime.now().year)
        return JourExceptionnel.objects.filter(annee=annee)


class CongeAnnuelViewSet(viewsets.ReadOnlyModelViewSet):
    """Périodes de congés annuels"""
    queryset = CongeAnnuel.objects.filter(est_actif=True)
    serializer_class = CongeAnnuelSerializer
    permission_classes = [permissions.IsAuthenticated]


class CongeViewSet(viewsets.ModelViewSet):
    """Gestion des demandes de congés"""
    serializer_class = CongeSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # Admin voit tout
        if user.is_staff or user.is_superuser:
            return Conge.objects.all()
        
        # Utilisateur normal voit seulement ses congés
        return Conge.objects.filter(utilisateur=user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return CongeCreateSerializer
        return CongeSerializer
    
    def perform_create(self, serializer):
        """Crée le congé SANS déduire le solde (déduction à la validation)"""
        user = self.request.user
        data = self.request.data
        
        # Si user_id fourni et que l'utilisateur peut gérer d'autres
        target_user_id = data.get('user_id')
        if target_user_id and self._can_manage_user(user, target_user_id):
            from users.models import User
            target_user = User.objects.get(id=target_user_id)
        else:
            target_user = user
        
        conge = serializer.save(utilisateur=target_user)
        
        # NE PAS déduire ici - on attend la validation
        # Juste calculer les jours déduits pour info
        conge.calculer_deduction()
        conge.save()
        
        return conge
    
    def _can_manage_user(self, manager, target_user_id):
        """Vérifie si le manager peut gérer l'utilisateur cible"""
        if manager.is_superuser:
            return True
        
        from users.models import User
        try:
            target = User.objects.get(id=target_user_id)
        except User.DoesNotExist:
            return False
        
        # Vérifier si dans les équipes gérées
        equipes_manager = set()
        for eq in manager.equipes_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
        
        for eq in manager.equipes_co_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
        
        return target.equipe_id in equipes_manager
    
    @action(detail=False, methods=['get'])
    def mes_conges(self, request):
        """Liste les congés de l'utilisateur connecté avec filtres"""
        # CORRECTION : Toujours filtrer par l'utilisateur connecté
        queryset = Conge.objects.filter(utilisateur=request.user)
        
        # Filtre par année
        annee = request.query_params.get('annee')
        if annee:
            queryset = queryset.filter(date_debut__year=annee)
        
        # Filtre par statut
        statut = request.query_params.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def calendrier(self, request):
        """Retourne les événements pour le calendrier selon les permissions"""
        annee = int(request.query_params.get('annee', datetime.now().year))
        pole_id = request.query_params.get('pole')
        equipe_id = request.query_params.get('equipe')
        # NOUVEAU : Paramètre optionnel pour filtrer par statut
        statut_filter = request.query_params.get('statut')
        
        user = request.user
        events = []
        
        # Déterminer le scope de visibilité
        is_superadmin = user.is_superuser
        is_admin = user.is_staff
        is_manager = user.equipes_gerees.exists() or user.equipes_co_gerees.exists()
        
        # MODIFICATION : Inclure tous les statuts sauf 'annule' par défaut
        # ou filtrer selon le paramètre
        statuts_inclus = ['en_attente', 'approuve', 'refuse']
        if statut_filter:
            statuts_inclus = [statut_filter]
        
        # Base queryset pour les congés
        conges_queryset = Conge.objects.filter(
            date_debut__year=annee,
            statut__in=statuts_inclus  # ← MODIFIÉ : inclure 'refuse'
        )
        
        if is_superadmin:
            # Super admin : voit tout, filtres optionnels
            if pole_id:
                conges_queryset = conges_queryset.filter(utilisateur__pole_id=pole_id)
            if equipe_id:
                conges_queryset = conges_queryset.filter(utilisateur__equipe_id=equipe_id)
                
        elif is_admin and is_manager:
            # Admin + Manager/Co-manager : voit son équipe et sous-équipes
            equipes_accessibles = user.sous_equipes_accessibles
            
            conges_queryset = conges_queryset.filter(
                utilisateur__equipe__in=equipes_accessibles
            )
            
            if pole_id:
                conges_queryset = conges_queryset.filter(
                    utilisateur__pole_id=pole_id,
                    utilisateur__equipe__in=equipes_accessibles
                )
            if equipe_id:
                if int(equipe_id) in [eq.id for eq in equipes_accessibles]:
                    conges_queryset = conges_queryset.filter(utilisateur__equipe_id=equipe_id)
                    
        elif is_manager:
            # Manager/Co-manager simple : même logique
            equipes_accessibles = user.sous_equipes_accessibles
            conges_queryset = conges_queryset.filter(
                utilisateur__equipe__in=equipes_accessibles
            )
            
            if equipe_id and int(equipe_id) in [eq.id for eq in equipes_accessibles]:
                conges_queryset = conges_queryset.filter(utilisateur__equipe_id=equipe_id)
                
        else:
            # Utilisateur normal : ne voit que ses congés
            conges_queryset = conges_queryset.filter(utilisateur=user)
        
        # Sérialisation des congés
        for conge in conges_queryset.select_related('utilisateur'):
            
            event = {
                'id': f"conge_{conge.id}",
                'title': f"{conge.utilisateur.username.upper()} - {conge.utilisateur.get_display_name().upper()} - {conge.get_type_conge_display()} - {'✔️' if conge.motif else ''}",
                'start': conge.date_debut.isoformat(),
                'allDay': True,
                'color': self._get_conge_color(conge.type_conge, conge.statut),
                'type': 'conge',
                'user_id': conge.utilisateur.id,
                'equipe_id': conge.utilisateur.equipe_id if conge.utilisateur.equipe else None,
                'statut': conge.statut  # ← AJOUTÉ
            }
            
            # Si date_fin > date_debut, ajouter 'end'
            if conge.date_fin > conge.date_debut:
                from datetime import timedelta
                day_after_end = conge.date_fin + timedelta(days=1)
                event['end'] = day_after_end.isoformat()
            
            events.append(event)
        
        # Jours fériés (tout le monde les voit)
        if is_superadmin or not equipe_id:
            jours_feries = JourFerie.objects.filter(est_actif=True)
            for jf in jours_feries:
                try:
                    date_jf = datetime(annee, jf.mois, jf.jour).date()
                    events.append({
                        'id': f"ferie_{jf.id}",
                        'title': f"🎉 {jf.nom}",
                        'start': date_jf.isoformat(),
                        'allDay': True,
                        'color': '#ffeb3b',
                        'type': 'ferie',
                        'isBlocked': True
                    })
                except ValueError:
                    continue
        
        # Jours exceptionnels (visibles selon scope)
        if is_superadmin:
            jours_exc = JourExceptionnel.objects.filter(annee=annee)
        elif is_manager or is_admin:
            jours_exc = JourExceptionnel.objects.filter(annee=annee)
        else:
            jours_exc = JourExceptionnel.objects.none()
            
        for je in jours_exc:
            events.append({
                'id': f"exc_{je.id}",
                'title': f"⛔ {je.description or 'Exceptionnel'}",
                'start': je.date.isoformat(),
                'allDay': True,
                'color': '#ff5722' if je.type_jour == 'exceptionnel' else '#4caf50',
                'type': 'exceptionnel',
                'isBlocked': je.type_jour == 'exceptionnel'
            })
        
        # Congés annuels
        if is_superadmin or is_manager or is_admin:
            conges_annuels = CongeAnnuel.objects.filter(annee=annee, est_actif=True)
            for ca in conges_annuels:
                events.append({
                    'id': f"annuel_{ca.id}",
                    'title': f"🏖️ {ca.nom}",
                    'start': ca.date_debut.isoformat(),
                    'end': ca.date_fin.isoformat(),
                    'allDay': True,
                    'color': '#9c27b0',
                    'type': 'conge_annuel',
                    'isBlocked': True
                })
        
        # Weekends (tout le monde)
        from datetime import timedelta
        current = datetime(annee, 1, 1).date()
        while current.year == annee:
            if current.weekday() >= 5:
                events.append({
                    'id': f"weekend_{current.isoformat()}",
                    'title': "📅 Weekend",
                    'start': current.isoformat(),
                    'allDay': True,
                    'color': '#e0e0e0',
                    'type': 'weekend',
                    'isBlocked': True
                })
            current += timedelta(days=1)
        
        return Response(events)

    @action(detail=False, methods=['get'])
    def utilisateurs_gerables(self, request):
        """Liste les utilisateurs pour lesquels le manager peut créer des congés"""
        from users.models import User
        
        user = request.user
        
        est_manager_principal = user.equipes_gerees.exists()
        est_co_manager = user.equipes_co_gerees.exists()
        est_manager = est_manager_principal or est_co_manager

        if user.is_superuser:
            users = User.objects.filter(is_active=True)
        elif user.is_staff and est_manager:
            equipes = user.sous_equipes_accessibles
            users = User.objects.filter(
                Q(equipe__in=equipes) | Q(id=user.id),
                is_active=True
            ).distinct()
        elif est_manager:
            equipes = user.sous_equipes_accessibles
            users = User.objects.filter(
                equipe__in=equipes,
                is_active=True
            ).distinct()
        else:
            users = User.objects.filter(id=user.id)
        
        data = [{
            'id': u.id,
            'display_name': u.get_display_name(),
            'username': u.username,
            'equipe_nom': u.equipe.nom if u.equipe else '',
            'solde_conge_actuelle': str(u.solde_conge_actuelle)
        } for u in users]
        
        return Response(data)
    
    def _get_conge_color(self, type_conge, statut='en_attente'):
        base_colors = {
            'matin': "#cf53ee",
            'midi': '#ffe0b2',
            'journee': '#bbdefb'
        }
        
        if statut == 'approuve':
            return '#4caf50'  # Vert validation
        
        # Si refusé : rouge
        if statut == 'refuse':
            return '#f44336'  # Rouge refus
        
        # Si annulé : gris
        if statut == 'annule':
            return '#9e9e9e'  # Gris
        
        # Par défaut (en_attente) : couleur du type
        return base_colors.get(type_conge, '#e3f2fd')

    
    
    @action(detail=True, methods=['post'])
    def valider(self, request, pk=None):
        """Valider une demande de congé ET déduire le solde"""
        conge = self.get_object()
        user = request.user
        
        if not self._can_manage_conge(user, conge):
            return Response({'error': 'Permission refusée'}, status=403)
        
        if conge.statut != 'en_attente':
            return Response({'error': 'Seuls les congés en attente peuvent être validés'}, status=400)
        
        from django.utils import timezone
        from datetime import timedelta
        
        config = TypeCongeConfig.objects.get(type_conge=conge.type_conge)
        dates_specifiques = request.data.get('dates', [])
        
        # ========== VALIDATION PARTIELLE ==========
        if dates_specifiques:
            # Trier les dates
            dates_specifiques.sort()
            jours_a_valider = 0
            
            # Calculer la déduction totale
            for date_str in dates_specifiques:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
                deduction = self._calculer_deduction_jour(date, conge.type_conge, config)
                jours_a_valider += deduction
            
            # Vérifier le solde
            if conge.utilisateur.solde_conge_actuelle < jours_a_valider:
                return Response({'error': 'Solde insuffisant'}, status=400)
            
            # DÉDUIRE le solde
            conge.utilisateur.consommer_conge(jours_a_valider)
            conge.utilisateur.save()
            
            # CRÉER LES NOUVELLES PÉRIODES
            nouvelles_periodes = []
            current_start = conge.date_debut
            dates_validees = set(dates_specifiques)
            
            # Parcourir toute la période
            current = conge.date_debut
            while current <= conge.date_fin:
                date_str = current.strftime('%Y-%m-%d')
                
                # Si cette date est validée et qu'on a une période en cours
                if date_str in dates_validees:
                    # Si on avait une période non validée avant, on la crée
                    if current_start < current:
                        nouvelles_periodes.append({
                            'debut': current_start,
                            'fin': current - timedelta(days=1),
                            'valide': False
                        })
                    
                    # Créer la journée validée (seule)
                    nouvelles_periodes.append({
                        'debut': current,
                        'fin': current,
                        'valide': True
                    })
                    
                    # Nouvelle période commence après
                    current_start = current + timedelta(days=1)
                
                current += timedelta(days=1)
            
            # Ajouter la dernière période si elle existe
            if current_start <= conge.date_fin:
                nouvelles_periodes.append({
                    'debut': current_start,
                    'fin': conge.date_fin,
                    'valide': False
                })
            
            # MARQUER L'ANCIEN CONGE COMME ANNULE (remplacé)
            conge.statut = 'remplace'
            conge.save()
            
            # CRÉER LES NOUVEAUX CONGES
            nouveaux_ids = []
            for periode in nouvelles_periodes:
                if periode['valide']:
                    # Congé validé directement
                    nouveau_conge = Conge.objects.create(
                        utilisateur=conge.utilisateur,
                        type_conge=conge.type_conge,
                        date_debut=periode['debut'],
                        date_fin=periode['fin'],
                        motif=conge.motif,
                        statut='approuve',
                        jours_deduits=self._calculer_deduction_periode(
                            periode['debut'], periode['fin'], conge.type_conge, config
                        ),
                        valide_par=user,
                        date_validation=timezone.now(),
                        conge_original=conge  # Lien vers l'original
                    )
                else:
                    # Congé en attente
                    nouveau_conge = Conge.objects.create(
                        utilisateur=conge.utilisateur,
                        type_conge=conge.type_conge,
                        date_debut=periode['debut'],
                        date_fin=periode['fin'],
                        motif=conge.motif,
                        statut='en_attente',
                        conge_original=conge  # Lien vers l'original
                    )
                    nouveau_conge.calculer_deduction()
                    nouveau_conge.save()
                
                nouveaux_ids.append(nouveau_conge.id)
            
            return Response({
                'success': True,
                'message': f'Congé fractionné en {len(nouvelles_periodes)} période(s)',
                'nouveaux_conges': nouveaux_ids,
                'ancien_conge_annule': conge.id
            })
        
        else:
            # Code existant pour validation complète
            if conge.utilisateur.solde_conge_actuelle < conge.jours_deduits:
                return Response({'error': 'Solde insuffisant'}, status=400)
            
            conge.utilisateur.consommer_conge(conge.jours_deduits)
            conge.utilisateur.save()
            
            conge.statut = 'approuve'
            conge.valide_par = user
            conge.date_validation = timezone.now()
            conge.save()
            
            return Response({'success': True, 'message': 'Congé validé'})
    
    @action(detail=True, methods=['post'])
    def refuser(self, request, pk=None):
        """Refuser une demande de congé (manager uniquement)"""
        conge = self.get_object()
        user = request.user
        
        if not self._can_manage_conge(user, conge):
            return Response({'error': 'Permission refusée'}, status=403)
        
        if conge.statut != 'en_attente':
            return Response({'error': 'Seuls les congés en attente peuvent être refusés'}, status=400)
        
        commentaire = request.data.get('commentaire', '')
        
        from django.utils import timezone
        
        # ✅ CORRECTION : Pour un refus, le solde n'a PAS été déduit (car en attente)
        # Donc on ne rembourse rien ! On vérifie juste au cas où...
        if conge.jours_deduits > 0:
            # Log pour debug, mais normalement ne devrait pas arriver
            print(f"Warning: Congé en attente avec jours_deduits > 0: {conge.id}")
        
        conge.statut = 'refuse'
        conge.refuse_par = user
        conge.date_refus = timezone.now()
        conge.commentaire_refus = commentaire
        conge.save()
        
        return Response({
            'success': True,
            'message': 'Congé refusé',
            'refuse_par': user.get_display_name(),
            'date_refus': conge.date_refus,
            'commentaire_refus': commentaire
        })
    
    def _can_manage_conge(self, manager, conge):
        """Vérifie si le manager peut gérer ce congé spécifique"""
        if manager.is_superuser:
            return True
        
        target_user = conge.utilisateur
        
        # Vérifier si dans les équipes gérées
        equipes_manager = set()
        for eq in manager.equipes_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
                # Sous-sous-équipes
                for sous_sous in sous.sous_equipes.all():
                    equipes_manager.add(sous_sous.id)
        
        for eq in manager.equipes_co_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
                for sous_sous in sous.sous_equipes.all():
                    equipes_manager.add(sous_sous.id)
        
        return target_user.equipe_id in equipes_manager
    
    
    
    @action(detail=True, methods=['post'])
    def annuler(self, request, pk=None):
        conge = self.get_object()
        
        if conge.utilisateur != request.user and not request.user.is_staff:
            return Response({'error': 'Permission refusée'}, status=403)
        
        if conge.statut == 'annule':
            return Response({'error': 'Déjà annulé'}, status=400)
        
        # ✅ GESTION DES CONGÉS FRACTIONNÉS
        if conge.statut == 'remplace':
            # Un congé fractionné a déjà été remplacé par d'autres
            # On ne peut pas l'annuler directement
            conges_fractionnes = Conge.objects.filter(conge_original=conge)
            if conges_fractionnes.exists():
                return Response({
                    'error': 'Ce congé a été fractionné. Veuillez annuler les périodes individuelles.',
                    'conges_fractionnes': [c.id for c in conges_fractionnes]
                }, status=400)
        
        # ✅ REMBOURSEMENT CONDITIONNEL
        remboursement_effectue = False
        if conge.statut == 'approuve' and conge.jours_deduits > 0:
            conge.utilisateur.solde_conge_actuelle += conge.jours_deduits
            conge.utilisateur.solde_conge_consomme -= conge.jours_deduits
            conge.utilisateur.save()
            remboursement_effectue = True
            message = 'Congé approuvé annulé et solde remboursé'
        else:
            message = 'Congé annulé (aucun remboursement nécessaire)'
        
        conge.statut = 'annule'
        conge.save()
        
        return Response({
            'success': True,
            'message': message,
            'remboursement_effectue': remboursement_effectue
        })
    
    
    def _calculer_deduction_jour(self, date, type_conge, config=None):
        """Calcule la déduction pour un jour spécifique"""
        if not config:
            config = TypeCongeConfig.objects.get(type_conge=type_conge)
        
        weekday = date.weekday()
        
        # Weekend
        if weekday >= 5:
            return 0
        
        # Jours fériés
        if JourFerie.objects.filter(mois=date.month, jour=date.day, est_actif=True).exists():
            return 0
        
        # Jours exceptionnels
        if JourExceptionnel.objects.filter(date=date, type_jour='exceptionnel').exists():
            return 0
        
        deduction = config.deduction_jours
        
        # Règle Vendredi
        if weekday == 4 and config.vendredi_deduction:
            deduction = config.vendredi_deduction
        
        # Règle Jeudi
        elif weekday == 3 and config.jeudi_deduction:
            from django.utils import timezone
            from datetime import timedelta
            vendredi = date + timedelta(days=1)
            if (JourExceptionnel.objects.filter(date=vendredi, type_jour='exceptionnel').exists() or
                JourFerie.objects.filter(mois=vendredi.month, jour=vendredi.day).exists()):
                deduction = config.jeudi_deduction
        
        return deduction

    def _calculer_deduction_periode(self, date_debut, date_fin, type_conge, config=None):
        """Calcule la déduction pour une période"""
        total = 0
        current = date_debut
        while current <= date_fin:
            from django.utils import timezone
            from datetime import timedelta
            total += self._calculer_deduction_jour(current, type_conge, config)
            current += timedelta(days=1)
        return total
    
    
    def _solde_deja_deduit(self, conge):
        """Vérifie si le solde a déjà été déduit pour ce congé"""
        return conge.statut in ['approuve', 'remplace'] or (
            conge.jours_valides and len(conge.jours_valides) > 0
        )

    @action(detail=True, methods=['post'])
    def annuler_partiel(self, request, pk=None):
        """Annuler partiellement un congé (rembourser certains jours)"""
        conge = self.get_object()
        user = request.user
        
        # Vérifier les permissions
        if conge.utilisateur != user and not user.is_staff and not self._can_manage_conge(user, conge):
            return Response({'error': 'Permission refusée'}, status=403)
        
        if conge.statut != 'approuve':
            return Response({'error': 'Seuls les congés approuvés peuvent être annulés partiellement'}, status=400)
        
        from django.utils import timezone
        from datetime import timedelta
        import json
        
        dates_specifiques = request.data.get('dates', [])
        if not dates_specifiques:
            return Response({'error': 'Aucune date sélectionnée'}, status=400)
        
        # Trier les dates
        dates_specifiques.sort()
        config = TypeCongeConfig.objects.get(type_conge=conge.type_conge)
        
        # Calculer le remboursement total
        remboursement_total = 0
        for date_str in dates_specifiques:
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            remboursement = self._calculer_deduction_jour(date, conge.type_conge, config)
            remboursement_total += remboursement
        
        # REMBOURSER le solde
        conge.utilisateur.solde_conge_actuelle += remboursement_total
        conge.utilisateur.solde_conge_consomme -= remboursement_total
        conge.utilisateur.save()
        
        # Retirer ces dates des jours validés
        jours_valides = conge.jours_valides or []
        jours_restants = [j for j in jours_valides if j not in dates_specifiques]
        conge.jours_valides = jours_restants
        conge.save()
        
        # CRÉER LES NOUVELLES PÉRIODES (comme pour la validation)
        nouvelles_periodes = []
        current_start = conge.date_debut
        dates_annulees = set(dates_specifiques)
        
        current = conge.date_debut
        while current <= conge.date_fin:
            date_str = current.strftime('%Y-%m-%d')
            
            if date_str in dates_annulees:
                if current_start < current:
                    nouvelles_periodes.append({
                        'debut': current_start,
                        'fin': current - timedelta(days=1),
                        'valide': True  # Ces jours restent validés
                    })
                
                # Jour annulé (sera recréé en attente)
                nouvelles_periodes.append({
                    'debut': current,
                    'fin': current,
                    'valide': False  # Devient en attente
                })
                
                current_start = current + timedelta(days=1)
            
            current += timedelta(days=1)
        
        if current_start <= conge.date_fin:
            nouvelles_periodes.append({
                'debut': current_start,
                'fin': conge.date_fin,
                'valide': True
            })
        
        # MARQUER L'ANCIEN CONGE COMME REMPLACE
        conge.statut = 'remplace'
        conge.save()
        
        # CRÉER LES NOUVEAUX CONGES
        nouveaux_ids = []
        for periode in nouvelles_periodes:
            if periode['valide']:
                nouveau_conge = Conge.objects.create(
                    utilisateur=conge.utilisateur,
                    type_conge=conge.type_conge,
                    date_debut=periode['debut'],
                    date_fin=periode['fin'],
                    motif=conge.motif,
                    statut='approuve',
                    jours_deduits=self._calculer_deduction_periode(
                        periode['debut'], periode['fin'], conge.type_conge, config
                    ),
                    jours_valides=[format(d, '%Y-%m-%d') for d in self._generer_jours_periode(
                        periode['debut'], periode['fin']
                    )],
                    conge_original=conge
                )
            else:
                nouveau_conge = Conge.objects.create(
                    utilisateur=conge.utilisateur,
                    type_conge=conge.type_conge,
                    date_debut=periode['debut'],
                    date_fin=periode['fin'],
                    motif=conge.motif,
                    statut='en_attente',
                    conge_original=conge
                )
                nouveau_conge.calculer_deduction()
                nouveau_conge.save()
            
            nouveaux_ids.append(nouveau_conge.id)
        
        return Response({
            'success': True,
            'message': f'Annulation partielle: {len(dates_specifiques)} jour(s) remboursé(s)',
            'remboursement': str(remboursement_total),
            'nouveaux_conges': nouveaux_ids,
            'ancien_conge_annule': conge.id
        })

    def _generer_jours_periode(self, debut, fin):
        """Génère la liste des jours entre deux dates"""
        from datetime import timedelta
        jours = []
        current = debut
        while current <= fin:
            if current.weekday() < 5:  # Pas weekend
                jours.append(current)
            current += timedelta(days=1)
        return jours


    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export Excel de tous les congés (admin uniquement)"""
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({'error': 'Permission refusée'}, status=403)
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Congés"
        
        headers = ['Utilisateur', 'Type', 'Date début', 'Date fin', 'Jours déduits', 
                   'Statut', 'Motif', 'Date création']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        conges = Conge.objects.all().select_related('utilisateur').order_by('-date_debut')
        
        for conge in conges:
            ws.append([
                conge.utilisateur.get_display_name(),
                conge.get_type_conge_display(),
                conge.date_debut.strftime('%d/%m/%Y'),
                conge.date_fin.strftime('%d/%m/%Y'),
                float(conge.jours_deduits),
                conge.get_statut_display(),
                conge.motif or '',
                conge.date_creation.strftime('%d/%m/%Y %H:%M')
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=conges_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def export_mine(self, request):
        """Export Excel des congés de l'utilisateur connecté"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        user = request.user
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mes Congés"
        
        ws.append([f"Export des congés - {user.get_display_name()}"])
        ws.append([f"Solde actuel: {user.solde_conge_actuelle} jours"])
        ws.append([f"Solde consommé: {user.solde_conge_consomme} jours"])
        ws.append([])
        
        headers = ['Type', 'Date début', 'Date fin', 'Jours déduits', 'Statut', 'Motif']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[5]:
            cell.fill = header_fill
            cell.font = header_font
        
        conges = Conge.objects.filter(utilisateur=user).order_by('-date_debut')
        
        for conge in conges:
            ws.append([
                conge.get_type_conge_display(),
                conge.date_debut.strftime('%d/%m/%Y'),
                conge.date_fin.strftime('%d/%m/%Y'),
                float(conge.jours_deduits),
                conge.get_statut_display(),
                conge.motif or ''
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=mes_conges_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response
    
    
    