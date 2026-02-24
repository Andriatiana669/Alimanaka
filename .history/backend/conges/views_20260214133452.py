from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from datetime import datetime
from decimal import Decimal  # ← AJOUTÉ

from .models import TypeCongeConfig, JourFerie, JourExceptionnel, CongeAnnuel, Conge
from .serializers import (
    TypeCongeConfigSerializer, JourFerieSerializer,
    JourExceptionnelSerializer, CongeAnnuelSerializer,
    CongeSerializer, CongeCreateSerializer
)


class TypeCongeConfigViewSet(viewsets.ReadOnlyModelViewSet):
    """Configuration des types de congés (lecture seule pour les utilisateurs)"""
    queryset = TypeCongeConfig.objects.all()
    serializer_class = TypeCongeConfigSerializer
    permission_classes = [permissions.IsAuthenticated]


class JourFerieViewSet(viewsets.ReadOnlyModelViewSet):
    """Jours fériés"""
    queryset = JourFerie.objects.filter(est_actif=True)
    serializer_class = JourFerieSerializer
    permission_classes = [permissions.IsAuthenticated]


class JourExceptionnelViewSet(viewsets.ReadOnlyModelViewSet):
    """Jours exceptionnels et non-exceptionnels"""
    serializer_class = JourExceptionnelSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        annee = self.request.query_params.get('annee', datetime.now().year)
        return JourExceptionnel.objects.filter(annee=annee)


class CongeAnnuelViewSet(viewsets.ReadOnlyModelViewSet):
    """Périodes de congés annuels"""
    queryset = CongeAnnuel.objects.filter(est_actif=True)
    serializer_class = CongeAnnuelSerializer
    permission_classes = [permissions.IsAuthenticated]


class CongeViewSet(viewsets.ModelViewSet):
    """Gestion des demandes de congés"""
    serializer_class = CongeSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        # Admin voit tout
        if user.is_staff or user.is_superuser:
            return Conge.objects.all()
        
        # Utilisateur normal voit seulement ses congés
        return Conge.objects.filter(utilisateur=user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return CongeCreateSerializer
        return CongeSerializer
    
    def perform_create(self, serializer):
        """Crée le congé et met à jour le solde de l'utilisateur"""
        conge = serializer.save(utilisateur=self.request.user)
        
        # Déduire du solde si approuvé automatiquement ou en attente
        if conge.statut in ['en_attente', 'approuve']:
            # ← CORRIGÉ: Passer directement le Decimal, pas float
            conge.utilisateur.consommer_conge(conge.jours_deduits)
            conge.utilisateur.motif_conge = conge.motif
            conge.utilisateur.save()
        
        return conge
    
    @action(detail=False, methods=['get'])
    def mes_conges(self, request):
        """Liste les congés de l'utilisateur connecté avec filtres"""
        queryset = self.get_queryset().filter(utilisateur=request.user)
        
        # Filtre par année
        annee = request.query_params.get('annee')
        if annee:
            queryset = queryset.filter(date_debut__year=annee)
        
        # Filtre par statut
        statut = request.query_params.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    
    # Modification des privilège (Ninho)

    @action(detail=False, methods=['get'])
    def calendrier(self, request):
        """Retourne les événements pour le calendrier selon les permissions"""
        annee = int(request.query_params.get('annee', datetime.now().year))
        pole_id = request.query_params.get('pole')
        equipe_id = request.query_params.get('equipe')
        
        user = request.user
        events = []
        
        # Déterminer le scope de visibilité
        is_superadmin = user.is_superuser
        is_admin = user.is_staff
        is_manager = user.equipes_gerees.exists() or user.equipes_co_gerees.exists()
        
        # Base queryset pour les congés
        conges_queryset = Conge.objects.filter(
            date_debut__year=annee,
            statut__in=['en_attente', 'approuve']
        )
        
        if is_superadmin:
            # Super admin : voit tout, filtres optionnels
            if pole_id:
                conges_queryset = conges_queryset.filter(utilisateur__pole_id=pole_id)
            if equipe_id:
                conges_queryset = conges_queryset.filter(utilisateur__equipe_id=equipe_id)
                
        elif is_admin and is_manager:
            # Admin + Manager/Co-manager : voit son équipe et sous-équipes (pas parent)
            equipes_accessibles = user.sous_equipes_accessibles
            # Exclure l'équipe parente si existe
            if user.equipe and user.equipe.equipe_parente:
                # L'utilisateur est dans une sous-équipe, ne pas remonter au parent
                pass
            
            conges_queryset = conges_queryset.filter(
                utilisateur__equipe__in=equipes_accessibles
            )
            
            # Filtres additionnels si dans son scope
            if pole_id:
                conges_queryset = conges_queryset.filter(
                    utilisateur__pole_id=pole_id,
                    utilisateur__equipe__in=equipes_accessibles
                )
            if equipe_id:
                if int(equipe_id) in [eq.id for eq in equipes_accessibles]:
                    conges_queryset = conges_queryset.filter(utilisateur__equipe_id=equipe_id)
                    
        elif is_manager:
            # Manager/Co-manager simple : même logique mais sans override admin
            equipes_accessibles = user.sous_equipes_accessibles
            conges_queryset = conges_queryset.filter(
                utilisateur__equipe__in=equipes_accessibles
            )
            
            if equipe_id and int(equipe_id) in [eq.id for eq in equipes_accessibles]:
                conges_queryset = conges_queryset.filter(utilisateur__equipe_id=equipe_id)
                
        else:
            # Utilisateur normal : ne voit que ses congés
            conges_queryset = conges_queryset.filter(utilisateur=user)
        
        # Sérialisation des congés
        for conge in conges_queryset.select_related('utilisateur'):
            events.append({
                'id': f"conge_{conge.id}",
                'title': f"{conge.utilisateur.get_display_name().upper()} - {conge.get_type_conge_display()} - {'✔️' if conge.motif else '❌'}",
                'start': conge.date_debut.isoformat(),
                'end': conge.date_fin.isoformat() if conge.date_fin != conge.date_debut else None,
                'allDay': True,
                'color': self._get_conge_color(conge.type_conge),
                'type': 'conge',
                'user_id': conge.utilisateur.id,
                'equipe_id': conge.utilisateur.equipe_id if conge.utilisateur.equipe else None
            })
        
        # Jours fériés (tout le monde les voit)
        if is_superadmin or not equipe_id:  # Les fériés sont globaux
            jours_feries = JourFerie.objects.filter(est_actif=True)
            for jf in jours_feries:
                try:
                    date_jf = datetime(annee, jf.mois, jf.jour).date()
                    events.append({
                        'id': f"ferie_{jf.id}",
                        'title': f"🎉 {jf.nom}",
                        'start': date_jf.isoformat(),
                        'allDay': True,
                        'color': '#ffeb3b',
                        'type': 'ferie',
                        'isBlocked': True
                    })
                except ValueError:
                    continue
        
        # Jours exceptionnels (visibles selon scope)
        if is_superadmin:
            jours_exc = JourExceptionnel.objects.filter(annee=annee)
        elif is_manager or is_admin:
            # Managers voient les exceptionnels de leur scope
            jours_exc = JourExceptionnel.objects.filter(annee=annee)
        else:
            jours_exc = JourExceptionnel.objects.none()
            
        for je in jours_exc:
            events.append({
                'id': f"exc_{je.id}",
                'title': f"⛔ {je.description or 'Exceptionnel'}",
                'start': je.date.isoformat(),
                'allDay': True,
                'color': '#ff5722' if je.type_jour == 'exceptionnel' else '#4caf50',
                'type': 'exceptionnel',
                'isBlocked': je.type_jour == 'exceptionnel'
            })
        
        # Congés annuels
        if is_superadmin or is_manager or is_admin:
            conges_annuels = CongeAnnuel.objects.filter(annee=annee, est_actif=True)
            for ca in conges_annuels:
                events.append({
                    'id': f"annuel_{ca.id}",
                    'title': f"🏖️ {ca.nom}",
                    'start': ca.date_debut.isoformat(),
                    'end': ca.date_fin.isoformat(),
                    'allDay': True,
                    'color': '#9c27b0',
                    'type': 'conge_annuel',
                    'isBlocked': True
                })
        
        # Weekends (tout le monde)
        from datetime import timedelta
        current = datetime(annee, 1, 1).date()
        while current.year == annee:
            if current.weekday() >= 5:
                events.append({
                    'id': f"weekend_{current.isoformat()}",
                    'title': "📅 Weekend",
                    'start': current.isoformat(),
                    'allDay': True,
                    'color': '#e0e0e0',
                    'type': 'weekend',
                    'isBlocked': True
                })
            current += timedelta(days=1)
        
        return Response(events)


    # Nouvelle action pour récupérer les utilisateurs gérables (pour le dropdown)
    @action(detail=False, methods=['get'])
    def utilisateurs_gerables(self, request):
        """Liste les utilisateurs pour lesquels le manager peut créer des congés"""
        user = request.user
        
        est_manager_principal = user.equipes_gerees.exists()
        est_co_manager = user.equipes_co_gerees.exists()
        est_manager = est_manager_principal or est_co_manager
    
        if user.is_superuser:
            users = user.objects.filter(is_active=True)
        elif user.is_staff and (user.equipes_gerees.exists() or user.equipes_co_gerees.exists()):
            equipes = user.sous_equipes_accessibles
            users = user.objects.filter(
                Q(equipe__in=equipes) | Q(id=user.id),
                is_active=True
            ).distinct()
        elif user.equipes_gerees.exists() or user.equipes_co_gerees.exists():
            equipes = user.sous_equipes_accessibles
            users = user.objects.filter(
                equipe__in=equipes,
                is_active=True
            ).distinct()
        else:
            users = user.objects.filter(id=user.id)
        
        data = [{
            'id': u.id,
            'display_name': u.get_display_name(),
            'username': u.username,
            'equipe_nom': u.equipe.nom if u.equipe else '',
            'solde_conge_actuelle': str(u.solde_conge_actuelle)
        } for u in users]
        
        return Response(data)
    
    def _get_conge_color(self, type_conge):
        colors = {
            'matin': '#c8e6c9',
            'midi': '#ffe0b2',
            'journee': '#bbdefb'
        }
        return colors.get(type_conge, '#e3f2fd')
    
    @action(detail=True, methods=['post'])
    def annuler(self, request, pk=None):
        """Annuler une demande de congé et rembourser le solde"""
        conge = self.get_object()
        
        if conge.utilisateur != request.user and not request.user.is_staff:
            return Response({'error': 'Permission refusée'}, status=403)
        
        if conge.statut == 'annule':
            return Response({'error': 'Déjà annulé'}, status=400)
        
        # Rembourser le solde - utiliser Decimal
        if conge.jours_deduits > 0:
            # ← CORRIGÉ: Utiliser Decimal pour les opérations
            conge.utilisateur.solde_conge_actuelle += conge.jours_deduits
            conge.utilisateur.solde_conge_consomme -= conge.jours_deduits
            conge.utilisateur.save()
        
        conge.statut = 'annule'
        conge.save()
        
        return Response({'success': True, 'message': 'Congé annulé et solde remboursé'})
    
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export Excel de tous les congés (admin uniquement)"""
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({'error': 'Permission refusée'}, status=403)
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        
        # Créer le workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Congés"
        
        # En-têtes
        headers = ['Utilisateur', 'Type', 'Date début', 'Date fin', 'Jours déduits', 
                   'Statut', 'Motif', 'Date création']
        ws.append(headers)
        
        # Style en-tête
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        conges = Conge.objects.all().select_related('utilisateur').order_by('-date_debut')
        
        for conge in conges:
            ws.append([
                conge.utilisateur.get_display_name(),
                conge.get_type_conge_display(),
                conge.date_debut.strftime('%d/%m/%Y'),
                conge.date_fin.strftime('%d/%m/%Y'),
                float(conge.jours_deduits),
                conge.get_statut_display(),
                conge.motif or '',
                conge.date_creation.strftime('%d/%m/%Y %H:%M')
            ])
        
        # Ajustement colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=conges_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response
    
    
    @action(detail=False, methods=['get'])
    def export_mine(self, request):
        """Export Excel des congés de l'utilisateur connecté"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        user = request.user
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mes Congés"
        
        # Info utilisateur en haut
        ws.append([f"Export des congés - {user.get_display_name()}"])
        ws.append([f"Solde actuel: {user.solde_conge_actuelle} jours"])
        ws.append([f"Solde consommé: {user.solde_conge_consomme} jours"])
        ws.append([])  # Ligne vide
        
        # En-têtes
        headers = ['Type', 'Date début', 'Date fin', 'Jours déduits', 'Statut', 'Motif']
        ws.append(headers)
        
        # Style
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[5]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Données
        conges = Conge.objects.filter(utilisateur=user).order_by('-date_debut')
        
        for conge in conges:
            ws.append([
                conge.get_type_conge_display(),
                conge.date_debut.strftime('%d/%m/%Y'),
                conge.date_fin.strftime('%d/%m/%Y'),
                float(conge.jours_deduits),
                conge.get_statut_display(),
                conge.motif or ''
            ])
        
        # Ajustement colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=mes_conges_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response