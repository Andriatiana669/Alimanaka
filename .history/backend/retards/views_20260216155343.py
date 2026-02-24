from django.shortcuts import render

# Create your views here.
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q
from datetime import datetime
from decimal import Decimal

from .models import TypeRetardConfig, Retard, Rattrapage
from .serializers import (
    TypeRetardConfigSerializer, RetardSerializer, RetardCreateSerializer,
    RattrapageSerializer, RattrapageCreateSerializer
)


class TypeRetardConfigViewSet(viewsets.ReadOnlyModelViewSet):
    """Configuration des types de retards"""
    queryset = TypeRetardConfig.objects.all()
    serializer_class = TypeRetardConfigSerializer
    permission_classes = [permissions.IsAuthenticated]


class RetardViewSet(viewsets.ModelViewSet):
    """Gestion des retards"""
    serializer_class = RetardSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        
        if user.is_staff or user.is_superuser:
            return Retard.objects.all()
        
        return Retard.objects.filter(utilisateur=user)
    
    def get_serializer_class(self):
        if self.action == 'create':
            return RetardCreateSerializer
        return RetardSerializer
    
    def perform_create(self, serializer):
        """Crée le retard avec calcul automatique"""
        user = self.request.user
        
        # Récupérer l'heure de début prévue depuis la config admin ou valeur par défaut
        # Tu peux remplacer ça par une config globale ou par utilisateur
        heure_debut_prevue = self._get_heure_debut_prevue(user)
        
        retard = serializer.save(
            utilisateur=user,
            heure_debut_prevue=heure_debut_prevue
        )
        
        # Calculer le retard
        retard.calculer_retard()
        retard.save()
        
        return retard
    
    def _get_heure_debut_prevue(self, user):
        """Récupère l'heure de début prévue depuis la config"""
        # TODO: À adapter selon ta logique admin
        # Par défaut 08:00, mais tu peux le stocker dans User ou une config globale
        from datetime import time
        return time(8, 0)
    
    def _can_manage_retard(self, manager, retard):
        """Vérifie si le manager peut gérer ce retard"""
        if manager.is_superuser:
            return True
        
        target_user = retard.utilisateur
        
        # Même logique que pour les congés
        equipes_manager = set()
        for eq in manager.equipes_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
                for sous_sous in sous.sous_equipes.all():
                    equipes_manager.add(sous_sous.id)
        
        for eq in manager.equipes_co_gerees.all():
            equipes_manager.add(eq.id)
            for sous in eq.sous_equipes.all():
                equipes_manager.add(sous.id)
                for sous_sous in sous.sous_equipes.all():
                    equipes_manager.add(sous_sous.id)
        
        return target_user.equipe_id in equipes_manager
    
    @action(detail=False, methods=['get'])
    def mes_retards(self, request):
        """Liste les retards de l'utilisateur connecté"""
        queryset = Retard.objects.filter(utilisateur=request.user)
        
        # Filtre par statut
        statut = request.query_params.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        # Filtre par année
        annee = request.query_params.get('annee')
        if annee:
            queryset = queryset.filter(date__year=annee)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def calendrier(self, request):
        """Retourne les retards pour le calendrier"""
        annee = int(request.query_params.get('annee', datetime.now().year))
        pole_id = request.query_params.get('pole')
        equipe_id = request.query_params.get('equipe')
        
        user = request.user
        events = []
        
        is_superadmin = user.is_superuser
        is_admin = user.is_staff
        is_manager = user.equipes_gerees.exists() or user.equipes_co_gerees.exists()
        
        # Base queryset - exclure les annulés par défaut
        retards_queryset = Retard.objects.filter(
            date__year=annee,
            statut__in=['en_attente', 'approuve']
        )
        
        if is_superadmin:
            if pole_id:
                retards_queryset = retards_queryset.filter(utilisateur__pole_id=pole_id)
            if equipe_id:
                retards_queryset = retards_queryset.filter(utilisateur__equipe_id=equipe_id)
                
        elif is_admin and is_manager:
            equipes_accessibles = user.sous_equipes_accessibles
            retards_queryset = retards_queryset.filter(
                utilisateur__equipe__in=equipes_accessibles
            )
            
            if pole_id:
                retards_queryset = retards_queryset.filter(
                    utilisateur__pole_id=pole_id,
                    utilisateur__equipe__in=equipes_accessibles
                )
            if equipe_id:
                if int(equipe_id) in [eq.id for eq in equipes_accessibles]:
                    retards_queryset = retards_queryset.filter(utilisateur__equipe_id=equipe_id)
                    
        elif is_manager:
            equipes_accessibles = user.sous_equipes_accessibles
            retards_queryset = retards_queryset.filter(
                utilisateur__equipe__in=equipes_accessibles
            )
            
            if equipe_id and int(equipe_id) in [eq.id for eq in equipes_accessibles]:
                retards_queryset = retards_queryset.filter(utilisateur__equipe_id=equipe_id)
                
        else:
            retards_queryset = retards_queryset.filter(utilisateur=user)
        
        # Sérialisation des retards pour le calendrier
        for retard in retards_queryset.select_related('utilisateur'):
            # Couleur selon statut et heures restantes
            color = self._get_retard_color(retard)
            
            event = {
                'id': f"retard_{retard.id}",
                'title': f"⏰ {retard.utilisateur.username.upper()} - {retard.minutes_retard}min ({retard.heures_restantes}h rest)",
                'start': retard.date.isoformat(),
                'allDay': True,
                'color': color,
                'type': 'retard',
                'user_id': retard.utilisateur.id,
                'equipe_id': retard.utilisateur.equipe_id if retard.utilisateur.equipe else None,
                'statut': retard.statut,
                'heures_restantes': str(retard.heures_restantes)
            }
            
            events.append(event)
        
        return Response(events)
    
    def _get_retard_color(self, retard):
        """Détermine la couleur selon le statut et l'urgence"""
        if retard.statut == 'approuve':
            return '#4caf50'  # Vert : tout rattrapé
        
        if retard.heures_restantes <= 0:
            return '#4caf50'  # Vert
        
        # En attente : couleur selon heures restantes
        if retard.heures_restantes > 2:
            return '#f44336'  # Rouge : urgent (> 2h)
        elif retard.heures_restantes > 1:
            return '#ff9800'  # Orange : moyen
        else:
            return '#ffeb3b'  # Jaune : léger
    
    @action(detail=True, methods=['post'])
    def rattraper(self, request, pk=None):
        """
        Action "Rattraper le retard" 
        Crée un session de rattrapage et met à jour les heures restantes
        """
        retard = self.get_object()
        user = request.user
        
        # Vérifier permissions
        if retard.utilisateur != user and not self._can_manage_retard(user, retard):
            return Response({'error': 'Permission refusée'}, status=403)
        
        if retard.statut == 'annule':
            return Response({'error': 'Ce retard est annulé'}, status=400)
        
        if retard.heures_restantes <= 0:
            return Response({'error': 'Toutes les heures ont déjà été rattrapées'}, status=400)
        
        # Validation des données
        serializer = RattrapageCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        
        data = serializer.validated_data
        
        # Créer le rattrapage
        rattrapage = Rattrapage(
            retard=retard,
            date_rattrapage=data['date_rattrapage'],
            heure_debut=data['heure_debut'],
            heure_fin=data['heure_fin'],
            commentaire=data.get('commentaire', ''),
            valide_par=user if self._can_manage_retard(user, retard) else None
        )
        
        # Calculer les heures rattrapées
        heures_rattrapees = rattrapage.calculer_heures()
        
        # Vérifier qu'on ne rattrape pas plus que nécessaire (optionnel)
        # if heures_rattrapees > retard.heures_restantes:
        #     return Response({
        #         'error': f'Vous ne pouvez rattraper que {retard.heures_restantes}h maximum'
        #     }, status=400)
        
        rattrapage.save()
        
        # Mettre à jour le retard
        heures_restantes = retard.rattraper_heures(heures_rattrapees)
        
        # Si c'est un manager qui valide, marquer comme approuvé si tout est rattrapé
        if retard.heures_restantes <= 0 and self._can_manage_retard(user, retard):
            retard.approuve_par = user
            retard.date_approbation = datetime.now()
            retard.statut = 'approuve'
            retard.save()
        
        return Response({
            'success': True,
            'message': f'{heures_rattrapees}h rattrapées. Reste: {heures_restantes}h',
            'heures_restantes': str(heures_restantes),
            'statut': retard.statut,
            'rattrapage': RattrapageSerializer(rattrapage).data
        })
    
    @action(detail=True, methods=['post'])
    def annuler(self, request, pk=None):
        """Annuler un retard"""
        retard = self.get_object()
        user = request.user
        
        if retard.utilisateur != user and not user.is_staff:
            return Response({'error': 'Permission refusée'}, status=403)
        
        if retard.statut == 'annule':
            return Response({'error': 'Déjà annulé'}, status=400)
        
        commentaire = request.data.get('commentaire', '')
        
        retard.statut = 'annule'
        retard.annule_par = user
        retard.date_annulation = datetime.now()
        retard.commentaire_annulation = commentaire
        retard.save()
        
        return Response({
            'success': True,
            'message': 'Retard annulé'
        })
    
    @action(detail=False, methods=['get'])
    def utilisateurs_gerables(self, request):
        """Liste les utilisateurs pour lesquels le manager peut déclarer des retards"""
        from users.models import User
        
        user = request.user
        
        est_manager_principal = user.equipes_gerees.exists()
        est_co_manager = user.equipes_co_gerees.exists()
        est_manager = est_manager_principal or est_co_manager

        if user.is_superuser:
            users = User.objects.filter(is_active=True)
        elif user.is_staff and est_manager:
            equipes = user.sous_equipes_accessibles
            users = User.objects.filter(
                Q(equipe__in=equipes) | Q(id=user.id),
                is_active=True
            ).distinct()
        elif est_manager:
            equipes = user.sous_equipes_accessibles
            users = User.objects.filter(
                equipe__in=equipes,
                is_active=True
            ).distinct()
        else:
            users = User.objects.filter(id=user.id)
        
        data = [{
            'id': u.id,
            'display_name': u.get_display_name(),
            'username': u.username,
            'equipe_nom': u.equipe.nom if u.equipe else '',
        } for u in users]
        
        return Response(data)
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export Excel des retards (admin)"""
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({'error': 'Permission refusée'}, status=403)
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Retards"
        
        headers = ['Utilisateur', 'Date', 'Heure prévue', 'Heure arrivée', 'Minutes retard',
                   'Heures à rattraper', 'Heures restantes', 'Statut', 'Justificatif']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        retards = Retard.objects.all().select_related('utilisateur').order_by('-date')
        
        for retard in retards:
            ws.append([
                retard.utilisateur.get_display_name(),
                retard.date.strftime('%d/%m/%Y'),
                retard.heure_debut_prevue.strftime('%H:%M'),
                retard.heure_arrivee_reelle.strftime('%H:%M'),
                retard.minutes_retard,
                float(retard.heures_a_rattraper),
                float(retard.heures_restantes),
                retard.get_statut_display(),
                retard.motif_retard or ''
            ])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=retards_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response